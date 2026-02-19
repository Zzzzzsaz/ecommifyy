from fastapi import FastAPI, APIRouter, HTTPException, Query
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
import httpx
import calendar
import io
from fastapi.responses import StreamingResponse

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ.get('MONGO_URL', '')
if not mongo_url:
    print("WARNING: MONGO_URL not set, using placeholder")
    mongo_url = "mongodb://localhost:27017"
    
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ.get('DB_NAME', 'ecommify_db')]

app = FastAPI()
api_router = APIRouter(prefix="/api")

# ===== CONSTANTS =====
USERS = {
    "2409": {"name": "Admin", "role": "admin", "pin": "2409"},
    "2609": {"name": "Kacper", "role": "kacper", "pin": "2609"},
    "2509": {"name": "Szymon", "role": "szymon", "pin": "2509"},
}

DEFAULT_SHOPS = [
    {"id": 1, "name": "ecom1", "color": "#6366f1"},
    {"id": 2, "name": "ecom2", "color": "#10b981"},
    {"id": 3, "name": "ecom3", "color": "#f59e0b"},
    {"id": 4, "name": "ecom4", "color": "#ec4899"},
]

MONTHS_PL = {1: "Styczen", 2: "Luty", 3: "Marzec", 4: "Kwiecien", 5: "Maj", 6: "Czerwiec", 7: "Lipiec", 8: "Sierpien", 9: "Wrzesien", 10: "Pazdziernik", 11: "Listopad", 12: "Grudzien"}

DEFAULT_APP_SETTINGS = {
    "target_revenue": 250000,
    "vat_rate": 23,
    "currency": "PLN",
    "profit_split": 2,
    "app_name": "Ecommify Campaign Calculator",
}

async def get_shops_list():
    shops = await db.shops.find({}, {"_id": 0}).sort("id", 1).to_list(100)
    if not shops:
        for s in DEFAULT_SHOPS:
            await db.shops.insert_one({**s, "is_active": True, "created_at": datetime.now(timezone.utc).isoformat()})
        shops = await db.shops.find({}, {"_id": 0}).sort("id", 1).to_list(100)
    return shops

async def get_shop_names():
    shops = await get_shops_list()
    return {s["id"]: s["name"] for s in shops}

async def get_app_settings():
    settings = await db.app_settings.find_one({"_key": "main"}, {"_id": 0})
    if not settings:
        settings = {**DEFAULT_APP_SETTINGS, "_key": "main"}
        await db.app_settings.insert_one(settings)
        settings.pop("_id", None)
    return settings

# ===== MODELS =====
class LoginRequest(BaseModel):
    pin: str

class IncomeCreate(BaseModel):
    amount: float
    date: str
    description: str = ""
    shop_id: int

class ExpenseCreate(BaseModel):
    amount: float
    date: str
    campaign_name: str = ""
    shop_id: int

class TaskCreate(BaseModel):
    title: str
    description: str = ""
    assigned_to: str = "oboje"
    due_date: Optional[str] = None
    created_by: str = "Admin"

class TaskUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    assigned_to: Optional[str] = None
    status: Optional[str] = None
    due_date: Optional[str] = None
    priority: Optional[str] = None

class IdeaCreate(BaseModel):
    title: str
    description: str = ""
    category: str = "inne"
    link: str = ""
    priority: bool = False
    created_by: str = "Admin"

class IdeaUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    category: Optional[str] = None
    link: Optional[str] = None
    priority: Optional[bool] = None

class ShopifyConfigCreate(BaseModel):
    shop_id: int
    store_url: str
    api_token: str

class TikTokConfigCreate(BaseModel):
    name: str
    advertiser_id: str
    access_token: str
    linked_shop_ids: List[int] = []

class TikTokConfigUpdate(BaseModel):
    name: Optional[str] = None
    advertiser_id: Optional[str] = None
    access_token: Optional[str] = None
    linked_shop_ids: Optional[List[int]] = None

class ChatMessage(BaseModel):
    shop_id: int
    message: str

# ===== AUTH =====
@api_router.post("/auth/login")
async def login(req: LoginRequest):
    user = USERS.get(req.pin)
    if not user:
        raise HTTPException(status_code=401, detail="Nieprawidlowy PIN")
    shops = await get_shops_list()
    settings = await get_app_settings()
    return {"user": user, "shops": shops, "settings": settings}

# ===== SHOPS CRUD =====
class ShopCreate(BaseModel):
    name: str
    color: str = "#6366f1"

class ShopUpdate(BaseModel):
    name: Optional[str] = None
    color: Optional[str] = None
    is_active: Optional[bool] = None

@api_router.get("/shops")
async def list_shops():
    return await get_shops_list()

@api_router.post("/shops")
async def create_shop(shop: ShopCreate):
    shops = await get_shops_list()
    max_id = max((s["id"] for s in shops), default=0)
    doc = {"id": max_id + 1, "name": shop.name, "color": shop.color, "is_active": True, "created_at": datetime.now(timezone.utc).isoformat()}
    await db.shops.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.put("/shops/{shop_id}")
async def update_shop(shop_id: int, update: ShopUpdate):
    upd = {k: v for k, v in update.dict().items() if v is not None}
    if upd:
        await db.shops.update_one({"id": shop_id}, {"$set": upd})
    shop = await db.shops.find_one({"id": shop_id}, {"_id": 0})
    if not shop:
        raise HTTPException(status_code=404, detail="Nie znaleziono sklepu")
    return shop

@api_router.delete("/shops/{shop_id}")
async def delete_shop(shop_id: int):
    await db.shops.delete_one({"id": shop_id})
    return {"status": "ok"}

# ===== APP SETTINGS =====
@api_router.get("/app-settings")
async def get_settings():
    return await get_app_settings()

@api_router.put("/app-settings")
async def update_settings(body: dict):
    body.pop("_key", None)
    body.pop("_id", None)
    await db.app_settings.update_one({"_key": "main"}, {"$set": body}, upsert=True)
    return await get_app_settings()

# ===== INCOMES =====
@api_router.post("/incomes")
async def create_income(income: IncomeCreate):
    doc = {
        "id": str(uuid.uuid4()),
        "amount": income.amount,
        "date": income.date,
        "description": income.description,
        "shop_id": income.shop_id,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.incomes.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.get("/incomes")
async def get_incomes(shop_id: int = Query(...), date: Optional[str] = None, year: Optional[int] = None, month: Optional[int] = None):
    query = {"shop_id": shop_id}
    if date:
        query["date"] = date
    elif year and month:
        prefix = f"{year}-{month:02d}"
        query["date"] = {"$regex": f"^{prefix}"}
    return await db.incomes.find(query, {"_id": 0}).to_list(10000)

@api_router.delete("/incomes/{income_id}")
async def delete_income(income_id: str):
    result = await db.incomes.delete_one({"id": income_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Nie znaleziono")
    return {"status": "ok"}

# ===== EXPENSES =====
@api_router.post("/expenses")
async def create_expense(expense: ExpenseCreate):
    doc = {
        "id": str(uuid.uuid4()),
        "amount": expense.amount,
        "date": expense.date,
        "campaign_name": expense.campaign_name,
        "shop_id": expense.shop_id,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.expenses.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.get("/expenses")
async def get_expenses(shop_id: int = Query(...), date: Optional[str] = None, year: Optional[int] = None, month: Optional[int] = None):
    query = {"shop_id": shop_id}
    if date:
        query["date"] = date
    elif year and month:
        prefix = f"{year}-{month:02d}"
        query["date"] = {"$regex": f"^{prefix}"}
    return await db.expenses.find(query, {"_id": 0}).to_list(10000)

@api_router.delete("/expenses/{expense_id}")
async def delete_expense(expense_id: str):
    result = await db.expenses.delete_one({"id": expense_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Nie znaleziono")
    return {"status": "ok"}

# ===== MONTHLY STATS =====
@api_router.get("/monthly-stats")
async def get_monthly_stats(shop_id: int = Query(...), year: int = Query(...), month: int = Query(...)):
    prefix = f"{year}-{month:02d}"
    incomes = await db.incomes.find({"shop_id": shop_id, "date": {"$regex": f"^{prefix}"}}, {"_id": 0}).to_list(10000)
    costs = await db.costs.find({"shop_id": shop_id, "date": {"$regex": f"^{prefix}"}}, {"_id": 0}).to_list(10000)
    custom_columns = await db.custom_columns.find({}, {"_id": 0}).to_list(100)
    app_s = await get_app_settings()
    split = max(app_s.get("profit_split", 2), 1)

    days_in_month = calendar.monthrange(year, month)[1]
    days = {}
    for d in range(1, days_in_month + 1):
        ds = f"{year}-{month:02d}-{d:02d}"
        days[ds] = {
            "date": ds, "income": 0, "netto": 0, "profit": 0, "profit_pp": 0,
            "tiktok_ads": 0, "meta_ads": 0, "google_ads": 0, "ads_total": 0,
            "zwroty": 0, "custom_costs": {}
        }
        for cc in custom_columns:
            days[ds]["custom_costs"][cc["name"]] = 0

    for inc in incomes:
        if inc["date"] in days:
            days[inc["date"]]["income"] += inc["amount"]
    
    for cost in costs:
        dt = cost["date"]
        if dt in days:
            cat = cost.get("category", "inne")
            amt = cost.get("amount", 0)
            if cat == "tiktok":
                days[dt]["tiktok_ads"] += amt
            elif cat == "meta":
                days[dt]["meta_ads"] += amt
            elif cat == "google":
                days[dt]["google_ads"] += amt
            elif cat == "zwroty":
                days[dt]["zwroty"] += amt
            else:
                if cat in days[dt]["custom_costs"]:
                    days[dt]["custom_costs"][cat] += amt

    total_income = 0
    total_tiktok = 0
    total_meta = 0
    total_google = 0
    total_zwroty = 0
    total_custom = {cc["name"]: 0 for cc in custom_columns}
    
    for day in days.values():
        day["netto"] = round(day["income"] * 0.77, 2)
        day["ads_total"] = round(day["tiktok_ads"] + day["meta_ads"] + day["google_ads"], 2)
        total_day_costs = day["ads_total"] + day["zwroty"]
        for cc in custom_columns:
            if cc["column_type"] == "expense":
                total_day_costs += day["custom_costs"].get(cc["name"], 0)
        day["profit"] = round(day["netto"] - total_day_costs, 2)
        day["profit_pp"] = round(day["profit"] / split, 2)
        
        total_income += day["income"]
        total_tiktok += day["tiktok_ads"]
        total_meta += day["meta_ads"]
        total_google += day["google_ads"]
        total_zwroty += day["zwroty"]
        for cc in custom_columns:
            total_custom[cc["name"]] = total_custom.get(cc["name"], 0) + day["custom_costs"].get(cc["name"], 0)

    total_netto = round(total_income * 0.77, 2)
    total_ads = round(total_tiktok + total_meta + total_google, 2)
    total_all_costs = total_ads + total_zwroty
    for cc in custom_columns:
        if cc["column_type"] == "expense":
            total_all_costs += total_custom.get(cc["name"], 0)
    total_profit = round(total_netto - total_all_costs, 2)
    roi = round((total_profit / total_all_costs * 100), 2) if total_all_costs > 0 else 0

    return {
        "shop_id": shop_id, "year": year, "month": month,
        "total_income": round(total_income, 2),
        "total_ads": total_ads,
        "total_tiktok": round(total_tiktok, 2),
        "total_meta": round(total_meta, 2),
        "total_google": round(total_google, 2),
        "total_zwroty": round(total_zwroty, 2),
        "total_custom": total_custom,
        "total_netto": total_netto,
        "total_profit": total_profit,
        "profit_per_person": round(total_profit / split, 2),
        "roi": roi,
        "custom_columns": custom_columns,
        "days": sorted(days.values(), key=lambda x: x["date"])
    }

# ===== COMBINED MONTHLY STATS =====
@api_router.get("/combined-monthly-stats")
async def get_combined_monthly_stats(year: int = Query(...), month: int = Query(...)):
    prefix = f"{year}-{month:02d}"
    incomes = await db.incomes.find({"date": {"$regex": f"^{prefix}"}}, {"_id": 0}).to_list(10000)
    costs = await db.costs.find({"date": {"$regex": f"^{prefix}"}}, {"_id": 0}).to_list(10000)
    custom_columns = await db.custom_columns.find({}, {"_id": 0}).to_list(100)
    shops = await get_shops_list()
    shop_ids = [s["id"] for s in shops]
    app_s = await get_app_settings()

    days_in_month = calendar.monthrange(year, month)[1]
    days = {}
    for d in range(1, days_in_month + 1):
        ds = f"{year}-{month:02d}-{d:02d}"
        days[ds] = {
            "date": ds, "income": 0,
            "tiktok_ads": 0, "meta_ads": 0, "google_ads": 0, "ads_total": 0,
            "zwroty": 0, "custom_costs": {cc["name"]: 0 for cc in custom_columns},
            "shops": [{
                "shop_id": i, "income": 0,
                "tiktok_ads": 0, "meta_ads": 0, "google_ads": 0, "ads_total": 0,
                "zwroty": 0, "custom_costs": {cc["name"]: 0 for cc in custom_columns}
            } for i in shop_ids]
        }

    for inc in incomes:
        dt = inc["date"]
        if dt in days:
            days[dt]["income"] += inc["amount"]
            sid = inc.get("shop_id", 1)
            for s in days[dt]["shops"]:
                if s["shop_id"] == sid:
                    s["income"] += inc["amount"]

    for cost in costs:
        dt = cost["date"]
        if dt in days:
            cat = cost.get("category", "inne")
            amt = cost.get("amount", 0)
            sid = cost.get("shop_id", 1)
            
            if cat == "tiktok":
                days[dt]["tiktok_ads"] += amt
            elif cat == "meta":
                days[dt]["meta_ads"] += amt
            elif cat == "google":
                days[dt]["google_ads"] += amt
            elif cat == "zwroty":
                days[dt]["zwroty"] += amt
            elif cat in days[dt]["custom_costs"]:
                days[dt]["custom_costs"][cat] += amt
            
            for s in days[dt]["shops"]:
                if s["shop_id"] == sid:
                    if cat == "tiktok":
                        s["tiktok_ads"] += amt
                    elif cat == "meta":
                        s["meta_ads"] += amt
                    elif cat == "google":
                        s["google_ads"] += amt
                    elif cat == "zwroty":
                        s["zwroty"] += amt
                    elif cat in s["custom_costs"]:
                        s["custom_costs"][cat] += amt

    total_income = 0
    total_tiktok = 0
    total_meta = 0
    total_google = 0
    total_zwroty = 0
    total_custom = {cc["name"]: 0 for cc in custom_columns}
    split = max(app_s.get("profit_split", 2), 1)
    
    for day in days.values():
        day["netto"] = round(day["income"] * 0.77, 2)
        day["ads_total"] = round(day["tiktok_ads"] + day["meta_ads"] + day["google_ads"], 2)
        total_day_costs = day["ads_total"] + day["zwroty"]
        for cc in custom_columns:
            if cc["column_type"] == "expense":
                total_day_costs += day["custom_costs"].get(cc["name"], 0)
        day["profit"] = round(day["netto"] - total_day_costs, 2)
        day["profit_pp"] = round(day["profit"] / split, 2)
        
        for s in day["shops"]:
            s["netto"] = round(s["income"] * 0.77, 2)
            s["ads_total"] = round(s["tiktok_ads"] + s["meta_ads"] + s["google_ads"], 2)
            s_costs = s["ads_total"] + s["zwroty"]
            for cc in custom_columns:
                if cc["column_type"] == "expense":
                    s_costs += s["custom_costs"].get(cc["name"], 0)
            s["profit"] = round(s["netto"] - s_costs, 2)
            s["profit_pp"] = round(s["profit"] / split, 2)
        
        total_income += day["income"]
        total_tiktok += day["tiktok_ads"]
        total_meta += day["meta_ads"]
        total_google += day["google_ads"]
        total_zwroty += day["zwroty"]
        for cc in custom_columns:
            total_custom[cc["name"]] = total_custom.get(cc["name"], 0) + day["custom_costs"].get(cc["name"], 0)

    total_netto = round(total_income * 0.77, 2)
    total_ads = round(total_tiktok + total_meta + total_google, 2)
    total_all_costs = total_ads + total_zwroty
    for cc in custom_columns:
        if cc["column_type"] == "expense":
            total_all_costs += total_custom.get(cc["name"], 0)
    total_profit = round(total_netto - total_all_costs, 2)
    roi = round((total_profit / total_all_costs * 100), 2) if total_all_costs > 0 else 0

    now_utc = datetime.now(timezone.utc)
    today_num = now_utc.day if (year == now_utc.year and month == now_utc.month) else days_in_month

    streak = 0
    for d in range(today_num, 0, -1):
        ds = f"{year}-{month:02d}-{d:02d}"
        if ds in days and days[ds]["profit"] > 0:
            streak += 1
        elif ds in days and (days[ds]["income"] > 0 or days[ds]["ads_total"] > 0):
            break
        else:
            break

    active = [d for d in days.values() if d["income"] > 0]
    best = max(active, key=lambda x: x["profit"])["date"] if active else None
    forecast = round((total_income / today_num) * days_in_month, 2) if total_income > 0 and today_num > 0 else 0

    target = app_s.get("target_revenue", 250000)

    return {
        "year": year, "month": month,
        "total_income": round(total_income, 2), "total_ads": total_ads,
        "total_tiktok": round(total_tiktok, 2),
        "total_meta": round(total_meta, 2),
        "total_google": round(total_google, 2),
        "total_zwroty": round(total_zwroty, 2),
        "total_custom": total_custom,
        "total_netto": total_netto, "total_profit": total_profit,
        "profit_per_person": round(total_profit / split, 2), "roi": roi,
        "target": target, "progress": round(min(total_income / max(target, 1) * 100, 100), 2) if total_income > 0 else 0,
        "streak": streak, "best_day": best, "forecast": forecast,
        "custom_columns": custom_columns,
        "days": sorted(days.values(), key=lambda x: x["date"]),
        "settings": {"target_revenue": target, "profit_split": split, "vat_rate": app_s.get("vat_rate", 23)}
    }

# ===== TASKS =====
@api_router.get("/tasks")
async def get_tasks():
    return await db.tasks.find({}, {"_id": 0}).to_list(1000)

@api_router.post("/tasks")
async def create_task(task: TaskCreate):
    doc = {
        "id": str(uuid.uuid4()),
        "title": task.title,
        "description": task.description,
        "assigned_to": task.assigned_to,
        "status": "todo",
        "due_date": task.due_date,
        "created_by": task.created_by,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    await db.tasks.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.put("/tasks/{task_id}")
async def update_task(task_id: str, update: TaskUpdate):
    update_dict = {k: v for k, v in update.model_dump().items() if v is not None}
    update_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.tasks.update_one({"id": task_id}, {"$set": update_dict})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Nie znaleziono")
    updated = await db.tasks.find_one({"id": task_id}, {"_id": 0})
    return updated

@api_router.delete("/tasks/{task_id}")
async def delete_task(task_id: str):
    result = await db.tasks.delete_one({"id": task_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Nie znaleziono")
    return {"status": "ok"}

# ===== IDEAS =====
@api_router.get("/ideas")
async def get_ideas():
    ideas = await db.ideas.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return ideas

@api_router.post("/ideas")
async def create_idea(idea: IdeaCreate):
    doc = {
        "id": str(uuid.uuid4()),
        "title": idea.title,
        "description": idea.description,
        "category": idea.category,
        "link": idea.link,
        "priority": idea.priority,
        "created_by": idea.created_by,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.ideas.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.put("/ideas/{idea_id}")
async def update_idea(idea_id: str, update: IdeaUpdate):
    update_data = {k: v for k, v in update.dict().items() if v is not None}
    if update_data:
        await db.ideas.update_one({"id": idea_id}, {"$set": update_data})
    idea = await db.ideas.find_one({"id": idea_id}, {"_id": 0})
    if not idea:
        raise HTTPException(status_code=404, detail="Nie znaleziono")
    return idea

@api_router.delete("/ideas/{idea_id}")
async def delete_idea(idea_id: str):
    result = await db.ideas.delete_one({"id": idea_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Nie znaleziono")
    return {"status": "ok"}

# ===== SHOPIFY CONFIGS =====
@api_router.get("/shopify-configs")
async def get_shopify_configs():
    return await db.shopify_configs.find({}, {"_id": 0}).to_list(100)

@api_router.post("/shopify-configs")
async def save_shopify_config(config: ShopifyConfigCreate):
    existing = await db.shopify_configs.find_one({"shop_id": config.shop_id})
    if existing:
        await db.shopify_configs.update_one(
            {"shop_id": config.shop_id},
            {"$set": {"store_url": config.store_url, "api_token": config.api_token, "is_active": True}}
        )
        return await db.shopify_configs.find_one({"shop_id": config.shop_id}, {"_id": 0})
    doc = {
        "id": str(uuid.uuid4()),
        "shop_id": config.shop_id,
        "store_url": config.store_url,
        "api_token": config.api_token,
        "is_active": True,
        "last_sync": None,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.shopify_configs.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.delete("/shopify-configs/{shop_id}")
async def delete_shopify_config(shop_id: int):
    result = await db.shopify_configs.delete_one({"shop_id": shop_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Nie znaleziono")
    return {"status": "ok"}

# ===== TIKTOK CONFIGS =====
@api_router.get("/tiktok-configs")
async def get_tiktok_configs():
    return await db.tiktok_configs.find({}, {"_id": 0}).to_list(100)

@api_router.post("/tiktok-configs")
async def create_tiktok_config(config: TikTokConfigCreate):
    doc = {
        "id": str(uuid.uuid4()),
        "name": config.name,
        "advertiser_id": config.advertiser_id,
        "access_token": config.access_token,
        "linked_shop_ids": config.linked_shop_ids,
        "is_active": True,
        "last_sync": None,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.tiktok_configs.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.put("/tiktok-configs/{config_id}")
async def update_tiktok_config(config_id: str, update: TikTokConfigUpdate):
    update_dict = {k: v for k, v in update.model_dump().items() if v is not None}
    result = await db.tiktok_configs.update_one({"id": config_id}, {"$set": update_dict})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Nie znaleziono")
    return await db.tiktok_configs.find_one({"id": config_id}, {"_id": 0})

@api_router.delete("/tiktok-configs/{config_id}")
async def delete_tiktok_config(config_id: str):
    result = await db.tiktok_configs.delete_one({"id": config_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Nie znaleziono")
    return {"status": "ok"}

# ===== SYNC HELPERS =====
async def _sync_shopify(shop_id: int, year: int, month: int):
    config = await db.shopify_configs.find_one({"shop_id": shop_id, "is_active": True}, {"_id": 0})
    if not config:
        raise HTTPException(status_code=404, detail="Brak konfiguracji Shopify")
    store_url = config["store_url"]
    api_token = config["api_token"]
    days_in_month = calendar.monthrange(year, month)[1]
    start_date = f"{year}-{month:02d}-01T00:00:00Z"
    end_date = f"{year}-{month:02d}-{days_in_month}T23:59:59Z"
    try:
        async with httpx.AsyncClient() as http:
            resp = await http.get(
                f"https://{store_url}/admin/api/2024-10/orders.json",
                params={"status": "any", "created_at_min": start_date, "created_at_max": end_date, "limit": 250, "financial_status": "paid"},
                headers={"X-Shopify-Access-Token": api_token},
                timeout=30
            )
            if resp.status_code != 200:
                return {"status": "error", "detail": f"Shopify HTTP {resp.status_code}"}
            orders = resp.json().get("orders", [])
            prefix = f"{year}-{month:02d}"
            await db.incomes.delete_many({"shop_id": shop_id, "date": {"$regex": f"^{prefix}"}, "description": {"$regex": "\\[Shopify\\]"}})
            daily = {}
            for o in orders:
                d = o["created_at"][:10]
                daily[d] = daily.get(d, 0) + float(o.get("total_price", 0))
            for ds, total in daily.items():
                doc = {"id": str(uuid.uuid4()), "amount": round(total, 2), "date": ds, "description": "[Shopify] Auto-sync", "shop_id": shop_id, "created_at": datetime.now(timezone.utc).isoformat()}
                await db.incomes.insert_one(doc)
            await db.shopify_configs.update_one({"shop_id": shop_id}, {"$set": {"last_sync": datetime.now(timezone.utc).isoformat()}})
            return {"status": "ok", "orders": len(orders), "days": len(daily)}
    except Exception as e:
        return {"status": "error", "detail": str(e)}

async def _sync_tiktok(config_id: str, year: int, month: int):
    config = await db.tiktok_configs.find_one({"id": config_id, "is_active": True}, {"_id": 0})
    if not config:
        raise HTTPException(status_code=404, detail="Brak konfiguracji TikTok")
    days_in_month = calendar.monthrange(year, month)[1]
    try:
        async with httpx.AsyncClient() as http:
            resp = await http.get(
                "https://business-api.tiktok.com/open_api/v1.3/report/integrated/get/",
                headers={"Access-Token": config["access_token"]},
                params={
                    "advertiser_id": config["advertiser_id"],
                    "report_type": "BASIC", "data_level": "AUCTION_ADVERTISER",
                    "dimensions": '["stat_time_day"]', "metrics": '["spend"]',
                    "start_date": f"{year}-{month:02d}-01",
                    "end_date": f"{year}-{month:02d}-{days_in_month:02d}",
                    "page": 1, "page_size": 31
                },
                timeout=30
            )
            if resp.status_code != 200:
                return {"status": "error", "detail": f"TikTok HTTP {resp.status_code}"}
            data = resp.json()
            if data.get("code") != 0:
                return {"status": "error", "detail": data.get("message", "Unknown")}
            rows = data.get("data", {}).get("list", [])
            linked = config.get("linked_shop_ids", [])
            prefix = f"{year}-{month:02d}"
            for sid in linked:
                await db.expenses.delete_many({"shop_id": sid, "date": {"$regex": f"^{prefix}"}, "campaign_name": {"$regex": f"\\[TikTok:{config['name']}\\]"}})
            count = 0
            for row in rows:
                ds = row.get("dimensions", {}).get("stat_time_day", "")[:10]
                spend = float(row.get("metrics", {}).get("spend", 0))
                if spend > 0 and linked:
                    per_shop = round(spend / len(linked), 2)
                    for sid in linked:
                        doc = {"id": str(uuid.uuid4()), "amount": per_shop, "date": ds, "campaign_name": f"[TikTok:{config['name']}] Auto-sync", "shop_id": sid, "created_at": datetime.now(timezone.utc).isoformat()}
                        await db.expenses.insert_one(doc)
                        count += 1
            await db.tiktok_configs.update_one({"id": config_id}, {"$set": {"last_sync": datetime.now(timezone.utc).isoformat()}})
            return {"status": "ok", "rows": len(rows), "entries": count}
    except Exception as e:
        return {"status": "error", "detail": str(e)}

@api_router.post("/sync/shopify/{shop_id}")
async def sync_shopify(shop_id: int, year: int = Query(...), month: int = Query(...)):
    return await _sync_shopify(shop_id, year, month)

@api_router.post("/sync/tiktok/{config_id}")
async def sync_tiktok(config_id: str, year: int = Query(...), month: int = Query(...)):
    return await _sync_tiktok(config_id, year, month)

@api_router.post("/sync/all")
async def sync_all(year: int = Query(...), month: int = Query(...)):
    results = {"shopify": [], "tiktok": []}
    for sc in await db.shopify_configs.find({"is_active": True}, {"_id": 0}).to_list(100):
        r = await _sync_shopify(sc["shop_id"], year, month)
        results["shopify"].append({"shop_id": sc["shop_id"], **r})
    for tc in await db.tiktok_configs.find({"is_active": True}, {"_id": 0}).to_list(100):
        r = await _sync_tiktok(tc["id"], year, month)
        results["tiktok"].append({"config_id": tc["id"], **r})
    return results

# ===== AI CHAT =====
@api_router.post("/chat")
async def chat_endpoint(msg: ChatMessage):
    # AI disabled - no OpenAI key
    return {"response": "AI chat jest wyłączony na tej instancji."}

    assistant_doc = {
        "id": str(uuid.uuid4()), "shop_id": msg.shop_id, "role": "assistant",
        "content": response, "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.chat_history.insert_one(assistant_doc)

    return {"response": response, "id": assistant_doc["id"]}

# ===== AI ASSISTANT WITH ACTIONS =====
class AssistantMessage(BaseModel):
    message: str

AVAILABLE_ACTIONS = """
Dostepne akcje (uzyj dokladnie tego formatu w odpowiedzi):
[ACTION:get_orders] - pokaz zamowienia
[ACTION:get_products] - pokaz produkty  
[ACTION:get_returns] - pokaz zwroty
[ACTION:get_fulfillment] - pokaz realizacje
[ACTION:get_stats:YYYY:MM] - pokaz statystyki np. [ACTION:get_stats:2025:12]
[ACTION:get_shops] - pokaz sklepy
[ACTION:get_reminders] - pokaz przypomnienia
[ACTION:delete_product:ID] - usun produkt
[ACTION:delete_order:ID] - usun zamowienie
[ACTION:delete_return:ID] - cofnij zwrot
[ACTION:delete_fulfillment:ID] - usun z realizacji
[ACTION:delete_reminder:ID] - usun przypomnienie
[ACTION:create_product:nazwa:cena:doplata:shop_id] - dodaj produkt
[ACTION:create_reminder:tytul:data] - dodaj przypomnienie (data: YYYY-MM-DD)
[ACTION:update_fulfillment:ID:status] - zmien status realizacji (waiting/reminder_sent/check_payment/to_ship/unpaid/archived)
"""

async def execute_action(action_str: str):
    """Parse and execute action from AI response"""
    try:
        parts = action_str.split(":")
        action = parts[0]
        
        if action == "get_orders":
            orders = await db.orders.find({}, {"_id": 0}).sort("date", -1).to_list(20)
            return {"action": "get_orders", "success": True, "data": orders, "message": f"Znaleziono {len(orders)} zamowien"}
        
        elif action == "get_products":
            products = await db.products.find({}, {"_id": 0}).to_list(50)
            return {"action": "get_products", "success": True, "data": products, "message": f"Znaleziono {len(products)} produktow"}
        
        elif action == "get_returns":
            returns = await db.returns.find({}, {"_id": 0}).to_list(50)
            return {"action": "get_returns", "success": True, "data": returns, "message": f"Znaleziono {len(returns)} zwrotow"}
        
        elif action == "get_fulfillment":
            items = await db.fulfillment.find({}, {"_id": 0}).to_list(50)
            return {"action": "get_fulfillment", "success": True, "data": items, "message": f"Znaleziono {len(items)} pozycji w realizacji"}
        
        elif action == "get_stats" and len(parts) >= 3:
            year, month = int(parts[1]), int(parts[2])
            prefix = f"{year}-{month:02d}"
            incomes = await db.incomes.find({"date": {"$regex": f"^{prefix}"}}, {"_id": 0}).to_list(10000)
            expenses = await db.expenses.find({"date": {"$regex": f"^{prefix}"}}, {"_id": 0}).to_list(10000)
            ti = sum(i["amount"] for i in incomes)
            te = sum(e["amount"] for e in expenses)
            return {"action": "get_stats", "success": True, "data": {"income": ti, "expenses": te, "netto": round(ti * 0.77, 2), "profit": round(ti * 0.77 - te, 2)}, "message": f"Statystyki {month}/{year}"}
        
        elif action == "get_shops":
            shops = await db.shops.find({}, {"_id": 0}).to_list(20)
            return {"action": "get_shops", "success": True, "data": shops, "message": f"Znaleziono {len(shops)} sklepow"}
        
        elif action == "get_reminders":
            reminders = await db.reminders.find({}, {"_id": 0}).to_list(50)
            return {"action": "get_reminders", "success": True, "data": reminders, "message": f"Znaleziono {len(reminders)} przypomn"}
        
        elif action == "delete_product" and len(parts) >= 2:
            pid = parts[1]
            product = await db.products.find_one({"id": pid}, {"_id": 0})
            if product:
                await db.products.delete_one({"id": pid})
                return {"action": "delete_product", "success": True, "message": f"Usunieto produkt: {product.get('name', pid)}"}
            return {"action": "delete_product", "success": False, "message": "Nie znaleziono produktu"}
        
        elif action == "delete_order" and len(parts) >= 2:
            oid = parts[1]
            order = await db.orders.find_one({"id": oid}, {"_id": 0})
            if order:
                await db.orders.delete_one({"id": oid})
                await db.returns.delete_many({"order_id": oid})
                await db.fulfillment.delete_many({"order_id": oid})
                await db.sales_records.delete_many({"order_id": oid})
                return {"action": "delete_order", "success": True, "message": f"Usunieto zamowienie: {order.get('order_number', oid)}"}
            return {"action": "delete_order", "success": False, "message": "Nie znaleziono zamowienia"}
        
        elif action == "delete_return" and len(parts) >= 2:
            rid = parts[1]
            ret = await db.returns.find_one({"id": rid}, {"_id": 0})
            if ret:
                if ret.get("order_id"):
                    await db.orders.update_one({"id": ret["order_id"]}, {"$set": {"status": "new"}})
                await db.returns.delete_one({"id": rid})
                return {"action": "delete_return", "success": True, "message": "Cofnieto zwrot, zamowienie przywrocone"}
            return {"action": "delete_return", "success": False, "message": "Nie znaleziono zwrotu"}
        
        elif action == "delete_fulfillment" and len(parts) >= 2:
            fid = parts[1]
            doc = await db.fulfillment.find_one({"id": fid}, {"_id": 0})
            if doc:
                if doc.get("order_id"):
                    await db.orders.update_one({"id": doc["order_id"]}, {"$set": {"status": "new"}})
                await db.fulfillment.delete_one({"id": fid})
                return {"action": "delete_fulfillment", "success": True, "message": f"Usunieto z realizacji: {doc.get('order_number', fid)}"}
            return {"action": "delete_fulfillment", "success": False, "message": "Nie znaleziono pozycji"}
        
        elif action == "delete_reminder" and len(parts) >= 2:
            rid = parts[1]
            reminder = await db.reminders.find_one({"id": rid}, {"_id": 0})
            if reminder:
                await db.reminders.delete_one({"id": rid})
                return {"action": "delete_reminder", "success": True, "message": f"Usunieto przypomnienie: {reminder.get('title', rid)}"}
            return {"action": "delete_reminder", "success": False, "message": "Nie znaleziono przypomnienia"}
        
        elif action == "create_product" and len(parts) >= 5:
            name, price, extra, shop_id = parts[1], float(parts[2]), float(parts[3]), int(parts[4])
            doc = {"id": str(uuid.uuid4()), "name": name, "sku": "", "price": price, "extra_payment": extra, "shop_id": shop_id, "category": "", "source": "ai", "created_at": datetime.now(timezone.utc).isoformat()}
            await db.products.insert_one(doc)
            return {"action": "create_product", "success": True, "message": f"Dodano produkt: {name} (cena: {price} zl, doplata: {extra} zl)"}
        
        elif action == "create_reminder" and len(parts) >= 3:
            title, date = parts[1], parts[2]
            doc = {"id": str(uuid.uuid4()), "title": title, "date": date, "time": None, "recurring": "none", "done": False, "created_by": "AI", "created_at": datetime.now(timezone.utc).isoformat()}
            await db.reminders.insert_one(doc)
            return {"action": "create_reminder", "success": True, "message": f"Dodano przypomnienie: {title} na {date}"}
        
        elif action == "update_fulfillment" and len(parts) >= 3:
            fid, status = parts[1], parts[2]
            result = await db.fulfillment.update_one({"id": fid}, {"$set": {"status": status}})
            if result.modified_count > 0:
                return {"action": "update_fulfillment", "success": True, "message": f"Zmieniono status na: {status}"}
            return {"action": "update_fulfillment", "success": False, "message": "Nie znaleziono pozycji lub status nie zmieniony"}
        
        return {"action": action, "success": False, "message": f"Nieznana akcja: {action}"}
    except Exception as e:
        return {"action": "error", "success": False, "message": str(e)}

@api_router.post("/ai-assistant")
async def ai_assistant_endpoint(msg: AssistantMessage):
    # AI disabled - no OpenAI key
    return {"response": "AI asystent jest wyłączony na tej instancji.", "actions": []}
    
    # Parse and execute actions from response
    import re
    action_pattern = r'\[ACTION:([^\]]+)\]'
    actions_found = re.findall(action_pattern, response)
    
    executed_actions = []
    for action_str in actions_found:
        result = await execute_action(action_str)
        executed_actions.append(result)
    
    # Clean response text (remove action tags for display)
    clean_response = re.sub(action_pattern, '', response).strip()
    if not clean_response:
        clean_response = "Wykonano!"
    
    # Save assistant response
    assistant_doc = {"id": str(uuid.uuid4()), "role": "assistant", "content": clean_response, "actions": executed_actions, "created_at": datetime.now(timezone.utc).isoformat()}
    await db.ai_assistant_history.insert_one(assistant_doc)
    
    return {"response": clean_response, "actions": executed_actions, "id": assistant_doc["id"]}

@api_router.get("/ai-assistant/history")
async def get_ai_assistant_history(limit: int = Query(50)):
    return await db.ai_assistant_history.find({}, {"_id": 0}).sort("created_at", 1).to_list(limit)

@api_router.delete("/ai-assistant/history")
async def clear_ai_assistant_history():
    await db.ai_assistant_history.delete_many({})
    return {"status": "ok"}

@api_router.get("/chat-history")
async def get_chat_history(shop_id: int = Query(...), limit: int = Query(50)):
    return await db.chat_history.find({"shop_id": shop_id}, {"_id": 0}).sort("created_at", 1).to_list(limit)

@api_router.delete("/chat-history")
async def clear_chat_history(shop_id: int = Query(...)):
    await db.chat_history.delete_many({"shop_id": shop_id})
    return {"status": "ok"}

# ===== HEALTH =====
@api_router.get("/")
async def root():
    return {"message": "Ecommify API running"}

# ===== REMINDERS =====
class ReminderCreate(BaseModel):
    title: str
    date: str
    time: Optional[str] = None
    recurring: str = "none"
    created_by: str = "Admin"

class ReminderUpdate(BaseModel):
    title: Optional[str] = None
    date: Optional[str] = None
    time: Optional[str] = None
    recurring: Optional[str] = None
    done: Optional[bool] = None

@api_router.get("/reminders")
async def get_reminders():
    return await db.reminders.find({}, {"_id": 0}).to_list(1000)

@api_router.post("/reminders")
async def create_reminder(r: ReminderCreate):
    doc = {"id": str(uuid.uuid4()), "title": r.title, "date": r.date, "time": r.time, "recurring": r.recurring, "done": False, "created_by": r.created_by, "created_at": datetime.now(timezone.utc).isoformat()}
    await db.reminders.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.put("/reminders/{rid}")
async def update_reminder(rid: str, update: ReminderUpdate):
    ud = {k: v for k, v in update.model_dump().items() if v is not None}
    result = await db.reminders.update_one({"id": rid}, {"$set": ud})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Nie znaleziono")
    return await db.reminders.find_one({"id": rid}, {"_id": 0})

@api_router.delete("/reminders/{rid}")
async def delete_reminder(rid: str):
    await db.reminders.delete_one({"id": rid})
    return {"status": "ok"}

# ===== ORDERS =====
class OrderCreate(BaseModel):
    order_number: str = ""
    customer_name: str = ""
    customer_email: str = ""
    customer_phone: str = ""
    shipping_address: str = ""
    shipping_method: str = ""
    parcel_locker: str = ""
    payment_method: str = ""
    payment_gateway: str = ""
    transaction_id: str = ""
    items: List[dict] = []
    total: float
    date: str
    shop_id: int
    status: str = "new"

@api_router.get("/orders")
async def get_orders(shop_id: Optional[int] = None, year: Optional[int] = None, month: Optional[int] = None):
    q = {}
    if shop_id and shop_id > 0: q["shop_id"] = shop_id
    if year and month: q["date"] = {"$regex": f"^{year}-{month:02d}"}
    return await db.orders.find(q, {"_id": 0}).sort("date", -1).to_list(10000)

@api_router.post("/orders")
async def create_order(order: OrderCreate):
    doc = {
        "id": str(uuid.uuid4()),
        "order_number": order.order_number or f"ORD-{str(uuid.uuid4())[:8].upper()}",
        "customer_name": order.customer_name,
        "customer_email": order.customer_email,
        "customer_phone": order.customer_phone,
        "shipping_address": order.shipping_address,
        "shipping_method": order.shipping_method,
        "parcel_locker": order.parcel_locker,
        "payment_method": order.payment_method,
        "payment_gateway": order.payment_gateway,
        "transaction_id": order.transaction_id,
        "items": order.items,
        "total": order.total,
        "date": order.date,
        "shop_id": order.shop_id,
        "status": order.status,
        "source": "manual",
        "receipt_id": None,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.orders.insert_one(doc)
    doc.pop("_id", None)
    # Auto-create fulfillment with calculated extra_payment from products
    extra = 0
    for it in doc.get("items", []):
        product_name = it.get("name", it.get("description", ""))
        qty = it.get("quantity", 1)
        product = await db.products.find_one({"name": product_name, "shop_id": doc["shop_id"]}, {"_id": 0})
        if not product:
            product = await db.products.find_one({"name": product_name}, {"_id": 0})
        if product and product.get("extra_payment", 0) > 0:
            extra += product["extra_payment"] * qty
    source_month = doc["date"][:7]
    fdoc = {
        "id": str(uuid.uuid4()), "order_id": doc["id"],
        "order_number": doc["order_number"], "customer_name": doc["customer_name"],
        "customer_email": doc.get("customer_email", ""), "customer_phone": doc.get("customer_phone", ""),
        "shipping_address": doc.get("shipping_address", ""), "items": doc.get("items", []),
        "total": doc["total"], "extra_payment": round(extra, 2), "extra_payment_paid": False,
        "source_month": source_month, "status": "waiting", "notes": "",
        "tracking_number": "", "reminder_sent_at": None, "payment_checked_at": None,
        "shipped_at": None, "shop_id": doc["shop_id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.fulfillment.insert_one(fdoc)
    # Auto-create sales records (ewidencja) from order items
    order_items = doc.get("items", [])
    if order_items:
        for it in order_items:
            qty = it.get("quantity", 1)
            price = it.get("price", 0)
            brutto = round(price * qty, 2)
            netto = round(brutto / 1.23, 2)
            sr_doc = {
                "id": str(uuid.uuid4()), "date": doc["date"],
                "order_number": doc.get("order_number", ""),
                "product_name": it.get("name", it.get("description", "Produkt")),
                "quantity": qty, "netto": netto, "vat_rate": 23,
                "vat_amount": round(brutto - netto, 2), "brutto": brutto,
                "payment_method": doc.get("payment_gateway", doc.get("payment_method", "")),
                "shop_id": doc.get("shop_id", 1), "order_id": doc["id"],
                "source": "order", "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.sales_records.insert_one(sr_doc)
    else:
        # Single entry for whole order if no items
        total = doc.get("total", 0)
        netto = round(total / 1.23, 2)
        sr_doc = {
            "id": str(uuid.uuid4()), "date": doc["date"],
            "order_number": doc.get("order_number", ""),
            "product_name": f"Zamowienie {doc.get('order_number', '')}",
            "quantity": 1, "netto": netto, "vat_rate": 23,
            "vat_amount": round(total - netto, 2), "brutto": total,
            "payment_method": doc.get("payment_gateway", doc.get("payment_method", "")),
            "shop_id": doc.get("shop_id", 1), "order_id": doc["id"],
            "source": "order", "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.sales_records.insert_one(sr_doc)
    return doc

@api_router.put("/orders/{oid}/status")
async def update_order_status(oid: str, status: str = Query(...)):
    result = await db.orders.update_one({"id": oid}, {"$set": {"status": status}})
    if result.matched_count == 0: raise HTTPException(status_code=404, detail="Nie znaleziono")
    return await db.orders.find_one({"id": oid}, {"_id": 0})

@api_router.delete("/orders/{oid}")
async def delete_order(oid: str):
    await db.orders.delete_one({"id": oid})
    await db.returns.delete_many({"order_id": oid})
    await db.fulfillment.delete_many({"order_id": oid})
    await db.sales_records.delete_many({"order_id": oid})
    return {"status": "ok"}

# ===== RETURNS (ZWROTY) =====
class ReturnCreate(BaseModel):
    order_id: str
    reason: str = ""
    refund_amount: Optional[float] = None

@api_router.get("/returns")
async def get_returns(year: Optional[int] = None, month: Optional[int] = None, shop_id: Optional[int] = None):
    q = {}
    if year and month: q["date"] = {"$regex": f"^{year}-{month:02d}"}
    if shop_id and shop_id > 0: q["shop_id"] = shop_id
    return await db.returns.find(q, {"_id": 0}).sort("date", -1).to_list(10000)

@api_router.post("/returns")
async def create_return(r: ReturnCreate):
    order = await db.orders.find_one({"id": r.order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Nie znaleziono zamowienia")
    doc = {
        "id": str(uuid.uuid4()),
        "order_id": r.order_id,
        "order_number": order.get("order_number", ""),
        "customer_name": order.get("customer_name", ""),
        "customer_email": order.get("customer_email", ""),
        "customer_phone": order.get("customer_phone", ""),
        "items": order.get("items", []),
        "total": order.get("total", 0),
        "refund_amount": r.refund_amount if r.refund_amount is not None else order.get("total", 0),
        "reason": r.reason,
        "date": order.get("date", ""),
        "shop_id": order.get("shop_id", 1),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.returns.insert_one(doc)
    doc.pop("_id", None)
    await db.orders.update_one({"id": r.order_id}, {"$set": {"status": "returned"}})
    return doc

@api_router.delete("/returns/{rid}")
async def delete_return(rid: str):
    ret = await db.returns.find_one({"id": rid}, {"_id": 0})
    if ret and ret.get("order_id"):
        await db.orders.update_one({"id": ret["order_id"]}, {"$set": {"status": "new"}})
    await db.returns.delete_one({"id": rid})
    return {"status": "ok"}

# ===== FULFILLMENT (REALIZACJA ZAMOWIEN) =====
# Statuses: waiting -> reminder_sent -> check_payment -> to_ship -> archived
# Also: unpaid (branched from check_payment)
class FulfillmentCreate(BaseModel):
    order_id: str
    extra_payment: float = 0
    notes: str = ""

class FulfillmentUpdate(BaseModel):
    status: Optional[str] = None
    extra_payment: Optional[float] = None
    extra_payment_paid: Optional[bool] = None
    notes: Optional[str] = None
    tracking_number: Optional[str] = None

@api_router.get("/fulfillment")
async def get_fulfillment(source_month: Optional[str] = None, status: Optional[str] = None):
    q = {}
    if source_month: q["source_month"] = source_month
    if status: q["status"] = status
    items = await db.fulfillment.find(q, {"_id": 0}).sort("created_at", -1).to_list(10000)
    now = datetime.now(timezone.utc)
    for item in items:
        if item["status"] == "reminder_sent" and item.get("reminder_sent_at"):
            sent_at = datetime.fromisoformat(item["reminder_sent_at"].replace("Z", "+00:00")) if isinstance(item["reminder_sent_at"], str) else item["reminder_sent_at"]
            if (now - sent_at).days >= 7:
                item["auto_check_ready"] = True
            else:
                days_left = 7 - (now - sent_at).days
                item["auto_check_ready"] = False
                item["days_until_check"] = max(days_left, 0)
    return items

@api_router.post("/fulfillment")
async def create_fulfillment(f: FulfillmentCreate):
    order = await db.orders.find_one({"id": f.order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Nie znaleziono zamowienia")
    existing = await db.fulfillment.find_one({"order_id": f.order_id}, {"_id": 0})
    if existing:
        upd = {}
        if f.extra_payment > 0: upd["extra_payment"] = f.extra_payment
        if f.notes: upd["notes"] = f.notes
        if upd: await db.fulfillment.update_one({"order_id": f.order_id}, {"$set": upd})
        updated = await db.fulfillment.find_one({"order_id": f.order_id}, {"_id": 0})
        return updated
    source_month = order.get("date", "")[:7]
    doc = {
        "id": str(uuid.uuid4()),
        "order_id": f.order_id,
        "order_number": order.get("order_number", ""),
        "customer_name": order.get("customer_name", ""),
        "customer_email": order.get("customer_email", ""),
        "customer_phone": order.get("customer_phone", ""),
        "shipping_address": order.get("shipping_address", ""),
        "items": order.get("items", []),
        "total": order.get("total", 0),
        "extra_payment": f.extra_payment,
        "extra_payment_paid": False,
        "source_month": source_month,
        "status": "waiting",
        "notes": f.notes,
        "tracking_number": "",
        "reminder_sent_at": None,
        "payment_checked_at": None,
        "shipped_at": None,
        "shop_id": order.get("shop_id", 1),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.fulfillment.insert_one(doc)
    doc.pop("_id", None)
    await db.orders.update_one({"id": f.order_id}, {"$set": {"status": "processing"}})
    return doc

@api_router.put("/fulfillment/{fid}")
async def update_fulfillment(fid: str, update: FulfillmentUpdate):
    upd = {}
    if update.status is not None:
        upd["status"] = update.status
        if update.status == "reminder_sent":
            upd["reminder_sent_at"] = datetime.now(timezone.utc).isoformat()
        elif update.status == "check_payment":
            upd["payment_checked_at"] = datetime.now(timezone.utc).isoformat()
        elif update.status == "archived":
            upd["shipped_at"] = datetime.now(timezone.utc).isoformat()
    if update.extra_payment is not None:
        upd["extra_payment"] = update.extra_payment
    if update.extra_payment_paid is not None:
        upd["extra_payment_paid"] = update.extra_payment_paid
    if update.notes is not None:
        upd["notes"] = update.notes
    if update.tracking_number is not None:
        upd["tracking_number"] = update.tracking_number
    if upd:
        await db.fulfillment.update_one({"id": fid}, {"$set": upd})
    doc = await db.fulfillment.find_one({"id": fid}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Nie znaleziono")
    if update.status == "archived" and doc.get("order_id"):
        await db.orders.update_one({"id": doc["order_id"]}, {"$set": {"status": "delivered"}})
    return doc

@api_router.post("/fulfillment/bulk-status")
async def bulk_update_fulfillment_status(source_month: str = Query(...), from_status: str = Query(...), to_status: str = Query(...)):
    upd = {"status": to_status}
    if to_status == "reminder_sent":
        upd["reminder_sent_at"] = datetime.now(timezone.utc).isoformat()
    elif to_status == "check_payment":
        upd["payment_checked_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.fulfillment.update_many({"source_month": source_month, "status": from_status}, {"$set": upd})
    return {"status": "ok", "updated": result.modified_count}

@api_router.delete("/fulfillment/{fid}")
async def delete_fulfillment(fid: str):
    doc = await db.fulfillment.find_one({"id": fid}, {"_id": 0})
    if doc and doc.get("order_id"):
        await db.orders.update_one({"id": doc["order_id"]}, {"$set": {"status": "new"}})
    await db.fulfillment.delete_one({"id": fid})
    return {"status": "ok"}

# ===== FULFILLMENT NOTES =====
@api_router.get("/fulfillment-notes")
async def get_fulfillment_notes(source_month: Optional[str] = None):
    q = {}
    if source_month: q["source_month"] = source_month
    return await db.fulfillment_notes.find(q, {"_id": 0}).sort("created_at", -1).to_list(1000)

@api_router.post("/fulfillment-notes")
async def create_fulfillment_note(body: dict):
    doc = {
        "id": str(uuid.uuid4()),
        "content": body.get("content", ""),
        "source_month": body.get("source_month", ""),
        "created_by": body.get("created_by", ""),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.fulfillment_notes.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.delete("/fulfillment-notes/{nid}")
async def delete_fulfillment_note(nid: str):
    await db.fulfillment_notes.delete_one({"id": nid})
    return {"status": "ok"}

@api_router.get("/fulfillment/reminder-check")
async def fulfillment_reminder_check():
    now = datetime.now(timezone.utc)
    current_day = now.day
    if now.month == 1:
        prev_month = f"{now.year - 1}-12"
    else:
        prev_month = f"{now.year}-{now.month - 1:02d}"
    waiting_count = await db.fulfillment.count_documents({"source_month": prev_month, "status": "waiting"})
    check_ready = 0
    reminder_items = await db.fulfillment.find({"status": "reminder_sent"}, {"_id": 0, "reminder_sent_at": 1}).to_list(10000)
    for item in reminder_items:
        if item.get("reminder_sent_at"):
            sent_at = datetime.fromisoformat(item["reminder_sent_at"].replace("Z", "+00:00")) if isinstance(item["reminder_sent_at"], str) else item["reminder_sent_at"]
            if (now - sent_at).days >= 7:
                check_ready += 1
    return {
        "is_15th": current_day >= 15,
        "prev_month": prev_month,
        "waiting_for_reminder": waiting_count,
        "ready_for_check": check_ready,
        "show_reminder": current_day >= 15 and waiting_count > 0,
    }

# ===== CATEGORIZED COSTS (TikTok, Meta, Google, Zwroty, Custom) =====
COST_CATEGORIES = ["tiktok", "meta", "google", "zwroty", "inne"]

class CostCreate(BaseModel):
    date: str
    shop_id: int
    category: str  # tiktok, meta, google, zwroty, inne, or custom column name
    amount: float
    description: str = ""

class CostUpdate(BaseModel):
    amount: Optional[float] = None
    description: Optional[str] = None

@api_router.get("/costs")
async def get_costs(shop_id: Optional[int] = None, date: Optional[str] = None, year: Optional[int] = None, month: Optional[int] = None, category: Optional[str] = None):
    q = {}
    if shop_id and shop_id > 0: q["shop_id"] = shop_id
    if date: q["date"] = date
    elif year and month: q["date"] = {"$regex": f"^{year}-{month:02d}"}
    if category: q["category"] = category
    return await db.costs.find(q, {"_id": 0}).sort("date", -1).to_list(10000)

@api_router.post("/costs")
async def create_cost(c: CostCreate):
    doc = {
        "id": str(uuid.uuid4()),
        "date": c.date,
        "shop_id": c.shop_id,
        "category": c.category,
        "amount": round(c.amount, 2),
        "description": c.description,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.costs.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.put("/costs/{cost_id}")
async def update_cost(cost_id: str, update: CostUpdate):
    upd = {k: v for k, v in update.dict().items() if v is not None}
    if upd: await db.costs.update_one({"id": cost_id}, {"$set": upd})
    c = await db.costs.find_one({"id": cost_id}, {"_id": 0})
    if not c: raise HTTPException(status_code=404, detail="Nie znaleziono")
    return c

@api_router.delete("/costs/{cost_id}")
async def delete_cost(cost_id: str):
    await db.costs.delete_one({"id": cost_id})
    return {"status": "ok"}

# ===== CUSTOM COLUMNS =====
class CustomColumnCreate(BaseModel):
    name: str
    column_type: str  # "income" or "expense"
    color: str = "#888888"

class CustomColumnUpdate(BaseModel):
    name: Optional[str] = None
    column_type: Optional[str] = None
    color: Optional[str] = None

@api_router.get("/custom-columns")
async def get_custom_columns():
    return await db.custom_columns.find({}, {"_id": 0}).sort("created_at", 1).to_list(100)

@api_router.post("/custom-columns")
async def create_custom_column(c: CustomColumnCreate):
    existing = await db.custom_columns.find_one({"name": c.name}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Kolumna o tej nazwie juz istnieje")
    doc = {
        "id": str(uuid.uuid4()),
        "name": c.name,
        "column_type": c.column_type,
        "color": c.color,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.custom_columns.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.put("/custom-columns/{col_id}")
async def update_custom_column(col_id: str, update: CustomColumnUpdate):
    upd = {k: v for k, v in update.dict().items() if v is not None}
    if upd: await db.custom_columns.update_one({"id": col_id}, {"$set": upd})
    c = await db.custom_columns.find_one({"id": col_id}, {"_id": 0})
    if not c: raise HTTPException(status_code=404, detail="Nie znaleziono")
    return c

@api_router.delete("/custom-columns/{col_id}")
async def delete_custom_column(col_id: str):
    col = await db.custom_columns.find_one({"id": col_id}, {"_id": 0})
    if col:
        await db.costs.delete_many({"category": col["name"]})
    await db.custom_columns.delete_one({"id": col_id})
    return {"status": "ok"}

# ===== PRODUCTS =====
class ProductCreate(BaseModel):
    name: str
    sku: str = ""
    price: float = 0
    extra_payment: float = 0
    shop_id: int
    category: str = ""
    description: str = ""

class ProductUpdate(BaseModel):
    name: Optional[str] = None
    sku: Optional[str] = None
    price: Optional[float] = None
    extra_payment: Optional[float] = None
    shop_id: Optional[int] = None
    category: Optional[str] = None
    description: Optional[str] = None

@api_router.get("/products")
async def get_products(shop_id: Optional[int] = None):
    q = {}
    if shop_id and shop_id > 0: q["shop_id"] = shop_id
    return await db.products.find(q, {"_id": 0}).sort("name", 1).to_list(10000)

@api_router.post("/products")
async def create_product(p: ProductCreate):
    doc = {
        "id": str(uuid.uuid4()),
        "name": p.name, "sku": p.sku, "price": round(p.price, 2),
        "extra_payment": round(p.extra_payment, 2), "shop_id": p.shop_id,
        "category": p.category, "description": p.description,
        "source": "manual", "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.products.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.put("/products/{pid}")
async def update_product(pid: str, update: ProductUpdate):
    upd = {k: v for k, v in update.dict().items() if v is not None}
    if upd: await db.products.update_one({"id": pid}, {"$set": upd})
    p = await db.products.find_one({"id": pid}, {"_id": 0})
    if not p: raise HTTPException(status_code=404, detail="Nie znaleziono")
    return p

@api_router.delete("/products/{pid}")
async def delete_product(pid: str):
    await db.products.delete_one({"id": pid})
    return {"status": "ok"}

# ===== COMPANY SETTINGS =====
class CompanySettings(BaseModel):
    name: str = ""
    nip: str = ""
    address: str = ""
    city: str = ""
    postal_code: str = ""
    bank_name: str = ""
    bank_account: str = ""
    email: str = ""
    phone: str = ""

@api_router.get("/company-settings")
async def get_company():
    doc = await db.company_settings.find_one({}, {"_id": 0})
    return doc or {}

@api_router.put("/company-settings")
async def update_company(s: CompanySettings):
    data = s.model_dump()
    await db.company_settings.update_one({}, {"$set": data}, upsert=True)
    return data

# ===== NOTES =====
class NoteCreate(BaseModel):
    date: str
    shop_id: int = 0
    content: str
    created_by: str = "Admin"

@api_router.get("/notes")
async def get_notes(date: Optional[str] = None, year: Optional[int] = None, month: Optional[int] = None):
    q = {}
    if date: q["date"] = date
    elif year and month: q["date"] = {"$regex": f"^{year}-{month:02d}"}
    return await db.notes.find(q, {"_id": 0}).to_list(1000)

@api_router.post("/notes")
async def create_note(note: NoteCreate):
    doc = {"id": str(uuid.uuid4()), "date": note.date, "shop_id": note.shop_id, "content": note.content, "created_by": note.created_by, "created_at": datetime.now(timezone.utc).isoformat()}
    await db.notes.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.delete("/notes/{nid}")
async def delete_note(nid: str):
    await db.notes.delete_one({"id": nid})
    return {"status": "ok"}

# ===== SALES RECORDS (EWIDENCJA SPRZEDAZY) =====
class SalesRecordCreate(BaseModel):
    date: str
    order_number: str = ""
    product_name: str
    quantity: int = 1
    netto: float
    vat_rate: int = 23
    brutto: float
    payment_method: str = ""
    shop_id: int

@api_router.get("/sales-records")
async def get_sales_records(shop_id: Optional[int] = None, year: Optional[int] = None, month: Optional[int] = None, date: Optional[str] = None):
    q = {}
    if shop_id and shop_id > 0: q["shop_id"] = shop_id
    if date: q["date"] = date
    elif year and month: q["date"] = {"$regex": f"^{year}-{month:02d}"}
    return await db.sales_records.find(q, {"_id": 0}).sort("date", -1).to_list(10000)

@api_router.post("/sales-records")
async def create_sales_record(r: SalesRecordCreate):
    doc = {
        "id": str(uuid.uuid4()),
        "date": r.date,
        "order_number": r.order_number,
        "product_name": r.product_name,
        "quantity": r.quantity,
        "netto": round(r.netto, 2),
        "vat_rate": r.vat_rate,
        "vat_amount": round(r.brutto - r.netto, 2),
        "brutto": round(r.brutto, 2),
        "payment_method": r.payment_method,
        "shop_id": r.shop_id,
        "order_id": None,
        "source": "manual",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.sales_records.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.delete("/sales-records/{sid}")
async def delete_sales_record(sid: str):
    await db.sales_records.delete_one({"id": sid})
    return {"status": "ok"}

@api_router.post("/sales-records/generate-from-orders")
async def generate_sales_from_orders(year: int = Query(...), month: int = Query(...), shop_id: Optional[int] = None):
    q = {"date": {"$regex": f"^{year}-{month:02d}"}}
    if shop_id and shop_id > 0: q["shop_id"] = shop_id
    orders = await db.orders.find(q, {"_id": 0}).to_list(10000)
    existing_order_ids = set()
    existing = await db.sales_records.find({"order_id": {"$ne": None}, "date": {"$regex": f"^{year}-{month:02d}"}}, {"_id": 0, "order_id": 1}).to_list(10000)
    for e in existing:
        if e.get("order_id"): existing_order_ids.add(e["order_id"])
    generated = 0
    for order in orders:
        if order["id"] in existing_order_ids:
            continue
        order_items = order.get("items", [])
        if order_items:
            for it in order_items:
                qty = it.get("quantity", 1)
                price = it.get("price", 0)
                brutto = round(price * qty, 2)
                netto = round(brutto / 1.23, 2)
                doc = {
                    "id": str(uuid.uuid4()), "date": order["date"],
                    "order_number": order.get("order_number", ""),
                    "product_name": it.get("name", it.get("description", "Produkt")),
                    "quantity": qty, "netto": netto, "vat_rate": 23,
                    "vat_amount": round(brutto - netto, 2), "brutto": brutto,
                    "payment_method": order.get("payment_gateway", order.get("payment_method", "")),
                    "shop_id": order.get("shop_id", 1), "order_id": order["id"],
                    "source": "order", "created_at": datetime.now(timezone.utc).isoformat()
                }
                await db.sales_records.insert_one(doc)
                generated += 1
        else:
            total = order.get("total", 0)
            netto = round(total / 1.23, 2)
            doc = {
                "id": str(uuid.uuid4()), "date": order["date"],
                "order_number": order.get("order_number", ""),
                "product_name": f"Zamowienie {order.get('order_number', '')}",
                "quantity": 1, "netto": netto, "vat_rate": 23,
                "vat_amount": round(total - netto, 2), "brutto": total,
                "payment_method": order.get("payment_gateway", order.get("payment_method", "")),
                "shop_id": order.get("shop_id", 1), "order_id": order["id"],
                "source": "order", "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.sales_records.insert_one(doc)
            generated += 1
    return {"status": "ok", "generated": generated, "message": f"Dodano {generated} wpisow do ewidencji"}

@api_router.get("/sales-records/pdf/daily")
async def sales_pdf_daily(date: str = Query(...), shop_id: Optional[int] = None):
    from fpdf import FPDF
    q = {"date": date}
    if shop_id and shop_id > 0: q["shop_id"] = shop_id
    recs = await db.sales_records.find(q, {"_id": 0}).sort("date", 1).to_list(10000)
    company = await db.company_settings.find_one({}, {"_id": 0}) or {}
    sn = await get_shop_names()
    pdf = FPDF()
    pdf.add_page("L")
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, "EWIDENCJA SPRZEDAZY - DZIENNA", ln=True, align="C")
    pdf.set_font("Helvetica", "", 11)
    pdf.cell(0, 6, f"Data: {date}", ln=True, align="C")
    if company.get("name"):
        pdf.set_font("Helvetica", "", 9)
        nip_str = f" | NIP: {company['nip']}" if company.get("nip") else ""
        pdf.cell(0, 5, f"{company['name']}{nip_str}", ln=True, align="C")
    pdf.ln(5)
    cols = [("Lp.", 10, "C"), ("Nr zam.", 30, "L"), ("Produkt", 65, "L"), ("Ilosc", 14, "C"), ("Netto", 28, "R"), ("VAT", 12, "C"), ("VAT kwota", 24, "R"), ("Brutto", 28, "R"), ("Platnosc", 30, "L"), ("Sklep", 20, "L")]
    pdf.set_font("Helvetica", "B", 7)
    pdf.set_fill_color(230, 230, 240)
    for c in cols: pdf.cell(c[1], 7, c[0], border=1, align=c[2], fill=True)
    pdf.ln()
    pdf.set_font("Helvetica", "", 7)
    tn = tv = tb = 0
    for i, r in enumerate(recs, 1):
        pdf.cell(10, 6, str(i), border=1, align="C")
        pdf.cell(30, 6, str(r.get("order_number", ""))[:18], border=1)
        pdf.cell(65, 6, str(r.get("product_name", ""))[:38], border=1)
        pdf.cell(14, 6, str(r.get("quantity", 1)), border=1, align="C")
        pdf.cell(28, 6, f"{r.get('netto', 0):.2f} zl", border=1, align="R")
        pdf.cell(12, 6, f"{r.get('vat_rate', 23)}%", border=1, align="C")
        pdf.cell(24, 6, f"{r.get('vat_amount', 0):.2f} zl", border=1, align="R")
        pdf.cell(28, 6, f"{r.get('brutto', 0):.2f} zl", border=1, align="R")
        pdf.cell(30, 6, str(r.get("payment_method", ""))[:18], border=1)
        pdf.cell(20, 6, sn.get(r.get("shop_id", 0), ""), border=1)
        pdf.ln()
        tn += r.get("netto", 0); tv += r.get("vat_amount", 0); tb += r.get("brutto", 0)
    pdf.set_font("Helvetica", "B", 8)
    pdf.set_fill_color(240, 240, 250)
    pdf.cell(119, 7, "RAZEM:", border=1, align="R", fill=True)
    pdf.cell(28, 7, f"{tn:.2f} zl", border=1, align="R", fill=True)
    pdf.cell(12, 7, "", border=1, fill=True)
    pdf.cell(24, 7, f"{tv:.2f} zl", border=1, align="R", fill=True)
    pdf.cell(28, 7, f"{tb:.2f} zl", border=1, align="R", fill=True)
    pdf.cell(50, 7, f"Pozycji: {len(recs)}", border=1, fill=True)
    pdf.ln(8)
    pdf.set_font("Helvetica", "", 7)
    pdf.cell(0, 4, f"Wygenerowano: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')} UTC", ln=True, align="R")
    buf = io.BytesIO(); pdf.output(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": f'attachment; filename="ewidencja_dzienna_{date}.pdf"'})

@api_router.get("/sales-records/pdf/monthly")
async def sales_pdf_monthly(year: int = Query(...), month: int = Query(...), shop_id: Optional[int] = None):
    from fpdf import FPDF
    q = {"date": {"$regex": f"^{year}-{month:02d}"}}
    if shop_id and shop_id > 0: q["shop_id"] = shop_id
    recs = await db.sales_records.find(q, {"_id": 0}).sort("date", 1).to_list(10000)
    company = await db.company_settings.find_one({}, {"_id": 0}) or {}
    sn = await get_shop_names()
    pdf = FPDF()
    pdf.add_page("L")
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, "EWIDENCJA SPRZEDAZY - MIESIECZNA", ln=True, align="C")
    pdf.set_font("Helvetica", "", 11)
    pdf.cell(0, 6, f"{MONTHS_PL.get(month, '')} {year}", ln=True, align="C")
    if company.get("name"):
        pdf.set_font("Helvetica", "", 9)
        nip_str = f" | NIP: {company['nip']}" if company.get("nip") else ""
        addr = f" | {company.get('address', '')} {company.get('postal_code', '')} {company.get('city', '')}".strip()
        pdf.cell(0, 5, f"{company['name']}{nip_str}{addr}", ln=True, align="C")
    pdf.ln(5)
    cols = [("Lp.", 10, "C"), ("Data", 20, "L"), ("Nr zam.", 28, "L"), ("Produkt", 55, "L"), ("Ilosc", 14, "C"), ("Netto", 26, "R"), ("VAT", 12, "C"), ("VAT kw.", 22, "R"), ("Brutto", 26, "R"), ("Platnosc", 28, "L"), ("Sklep", 20, "L")]
    pdf.set_font("Helvetica", "B", 7)
    pdf.set_fill_color(230, 230, 240)
    for c in cols: pdf.cell(c[1], 7, c[0], border=1, align=c[2], fill=True)
    pdf.ln()
    pdf.set_font("Helvetica", "", 6)
    tn = tv = tb = 0
    for i, r in enumerate(recs, 1):
        pdf.cell(10, 5.5, str(i), border=1, align="C")
        pdf.cell(20, 5.5, r.get("date", ""), border=1)
        pdf.cell(28, 5.5, str(r.get("order_number", ""))[:16], border=1)
        pdf.cell(55, 5.5, str(r.get("product_name", ""))[:32], border=1)
        pdf.cell(14, 5.5, str(r.get("quantity", 1)), border=1, align="C")
        pdf.cell(26, 5.5, f"{r.get('netto', 0):.2f}", border=1, align="R")
        pdf.cell(12, 5.5, f"{r.get('vat_rate', 23)}%", border=1, align="C")
        pdf.cell(22, 5.5, f"{r.get('vat_amount', 0):.2f}", border=1, align="R")
        pdf.cell(26, 5.5, f"{r.get('brutto', 0):.2f}", border=1, align="R")
        pdf.cell(28, 5.5, str(r.get("payment_method", ""))[:16], border=1)
        pdf.cell(20, 5.5, sn.get(r.get("shop_id", 0), ""), border=1)
        pdf.ln()
        tn += r.get("netto", 0); tv += r.get("vat_amount", 0); tb += r.get("brutto", 0)
    pdf.set_font("Helvetica", "B", 8)
    pdf.set_fill_color(240, 240, 250)
    pdf.cell(127, 7, "RAZEM:", border=1, align="R", fill=True)
    pdf.cell(26, 7, f"{tn:.2f} zl", border=1, align="R", fill=True)
    pdf.cell(12, 7, "", border=1, fill=True)
    pdf.cell(22, 7, f"{tv:.2f} zl", border=1, align="R", fill=True)
    pdf.cell(26, 7, f"{tb:.2f} zl", border=1, align="R", fill=True)
    pdf.cell(48, 7, f"Pozycji: {len(recs)}", border=1, fill=True)
    pdf.ln(8)
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 5, f"Podsumowanie: Netto: {tn:.2f} zl | VAT: {tv:.2f} zl | Brutto: {tb:.2f} zl | Pozycji: {len(recs)}", ln=True)
    pdf.set_font("Helvetica", "", 7)
    pdf.cell(0, 4, f"Wygenerowano: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')} UTC", ln=True, align="R")
    buf = io.BytesIO(); pdf.output(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": f'attachment; filename="ewidencja_{year}_{month:02d}.pdf"'})

# ===== WEEKLY STATS =====
@api_router.get("/weekly-stats")
async def get_weekly_stats():
    now = datetime.now(timezone.utc)
    monday = now - timedelta(days=now.weekday())
    m_str = monday.strftime("%Y-%m-%d")
    s_str = (monday + timedelta(days=6)).strftime("%Y-%m-%d")
    pm_str = (monday - timedelta(days=7)).strftime("%Y-%m-%d")
    ps_str = (monday - timedelta(days=1)).strftime("%Y-%m-%d")
    async def wk(start, end):
        inc = await db.incomes.find({"date": {"$gte": start, "$lte": end}}, {"_id": 0}).to_list(10000)
        exp = await db.expenses.find({"date": {"$gte": start, "$lte": end}}, {"_id": 0}).to_list(10000)
        ti = sum(i["amount"] for i in inc); te = sum(e["amount"] for e in exp)
        n = round(ti * 0.77, 2); p = round(n - te, 2)
        return {"income": round(ti, 2), "ads": round(te, 2), "profit": p, "profit_pp": round(p / 2, 2)}
    cur = await wk(m_str, s_str)
    prev = await wk(pm_str, ps_str)
    ic = round(((cur["income"] - prev["income"]) / prev["income"] * 100) if prev["income"] > 0 else 0, 1)
    pc = round(((cur["profit"] - prev["profit"]) / abs(prev["profit"]) * 100) if prev["profit"] != 0 else 0, 1)
    return {"current": {"start": m_str, "end": s_str, **cur}, "previous": {"start": pm_str, "end": ps_str, **prev}, "income_change": ic, "profit_change": pc}

# ===== EXPORT EXCEL =====
@api_router.get("/export/excel")
async def export_excel(year: int = Query(...), month: int = Query(...), shop_id: Optional[int] = None):
    from openpyxl import Workbook
    prefix = f"{year}-{month:02d}"
    iq = {"date": {"$regex": f"^{prefix}"}}; eq = {"date": {"$regex": f"^{prefix}"}}
    if shop_id and shop_id > 0: iq["shop_id"] = shop_id; eq["shop_id"] = shop_id
    incs = await db.incomes.find(iq, {"_id": 0}).sort("date", 1).to_list(10000)
    exps = await db.expenses.find(eq, {"_id": 0}).sort("date", 1).to_list(10000)
    sn = await get_shop_names()
    wb = Workbook()
    ws1 = wb.active; ws1.title = "Przychody"
    ws1.append(["Data", "Sklep", "Kwota", "Opis"])
    for i in incs: ws1.append([i["date"], sn.get(i.get("shop_id", 0), ""), i["amount"], i.get("description", "")])
    ws2 = wb.create_sheet("Koszty Ads")
    ws2.append(["Data", "Sklep", "Kwota", "Kampania"])
    for e in exps: ws2.append([e["date"], sn.get(e.get("shop_id", 0), ""), e["amount"], e.get("campaign_name", "")])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="ecommify_{year}_{month:02d}.xlsx"'})

# ===== INCOME/EXPENSE DETAILS =====
@api_router.get("/incomes/details")
async def income_details(shop_id: int = Query(...), date: str = Query(...)):
    return await db.incomes.find({"shop_id": shop_id, "date": date}, {"_id": 0}).to_list(1000)

@api_router.get("/expenses/details")
async def expense_details(shop_id: int = Query(...), date: str = Query(...)):
    return await db.expenses.find({"shop_id": shop_id, "date": date}, {"_id": 0}).to_list(1000)

# ===== SETUP =====
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
